import csv
import json
import os
import tkinter as tk
from datetime import datetime
from tkinter import filedialog, messagebox, ttk
import threading
import matplotlib.pyplot as plt
import matplotlib.patches as patches
import cv2
from matplotlib.figure import Figure
from scipy.ndimage import uniform_filter1d
from scipy.signal import find_peaks
import numpy as np
import re
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


GRIDQUANT_VERSION = "gridquant_V2_0.py"  # For CSV file updates, this should match the version of the script that generated the CSV file.
LAST_CONFIG_FILE = "last_config.json"

def load_presets(filename="presets.json"):
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)

    with open(path, "r", encoding="utf-8") as preset_file:
        presets = json.load(preset_file)

    if not isinstance(presets, dict) or not presets:
        raise ValueError("presets.json must contain at least one preset.")
    return presets

def get_channel_name(channel, channel_names):

    if channel <= len(channel_names):
        return channel_names[channel-1]

    return f"Channel {channel}"

def draw_circles_and_display(result):
    image = result.brightfield_image
    filename = os.path.basename(result.file_path)

    if result.snaked_circles is not None: 
        X = result.snaked_circles[0]
        Y = result.snaked_circles[1]
        R = result.snaked_circles[2]

        iterator = zip(X, Y, R)

    else: 
        iterator = result.circles

    # Determine the scaling factor
    max_dimension = 800
    h, w = image.shape[:2]
    scale = min(max_dimension / h, max_dimension / w)

    # Resize the image to fit within the 800x800 box
    resized_image = cv2.resize(image, (int(w * scale), int(h * scale)))

    # Create a figure and axis with matplotlib
    fig, ax = plt.subplots()
    ax.imshow(resized_image, cmap='gray')

    # Draw the circles and labels
    for x, y, r in iterator:
        original_x = int(x)
        original_y = int(y)
        original_r = int(r)

        # Scale the circle positions and radius
        x = int(x * scale)
        y = int(y * scale)
        r = int(r * scale)

        # Draw the circle as a patch
        circle_patch = patches.Circle((x, y), r, edgecolor='red', facecolor='none', linewidth=2)
        ax.add_patch(circle_patch)

        # Draw the centroid
        ax.plot(x, y, 'go', markersize=5)  # Green dot for centroid
        # Label the centroid with its coordinates
        label = f"({original_x},{original_y},{original_r})"
        ax.text(x + 11, y - r, label, color='red', fontsize=8, ha='right')

    # Set the aspect of the plot to be equal
    ax.set_aspect('equal')

    # Add grid lines
    ax.grid(True)

    # Display the result with interactive zoom
    ax.set_title(f"{filename}\nDetected Circles")

    return fig

def plot_outliers(result):

    data = result.outlier_plot_data
    filename = os.path.basename(result.file_path)

    X = data["X"]
    Y = data["Y"]
    X_final = data["X_final"]
    Y_final = data["Y_final"]
    to_remove = data["to_remove"]
    nearest_distances = data["nearest_distances"]

    fig, ax = plt.subplots(figsize=(10,8))

    ax.plot(X_final, Y_final, "go", label="Valid Points")

    if len(to_remove):
        ax.plot(X[to_remove], Y[to_remove], "ro", label="Removed Points")

    for i in range(len(X)):
        ax.text(
            X[i],
            Y[i],
            f"{round(nearest_distances[i])}",
            fontsize=9,
            ha="right",
            color="blue"
        )

    ax.set_xlabel("X Coordinate")
    ax.set_ylabel("Y Coordinate")
    ax.set_title(f"{filename}\nCircle Centers with Nearest Neighbor Distances")
    ax.legend()
    ax.grid(True)

    return fig

def plot_missing(result):
    data = result.missing_plot_data
    filename = os.path.basename(result.file_path)
    
    X = data["X"]
    Y = data["Y"]
    new_circles = data["new_circles"]

    # Plot the result
    fig, ax = plt.subplots(figsize=(10, 8))
    ax.plot(X, Y, 'go', label='Original Circles')
    if new_circles:
        ax.plot([c[0] for c in new_circles], [c[1] for c in new_circles], 'bo', label='Added Circles')
    
    ax.set_xlabel('X Coordinate')
    ax.set_ylabel('Y Coordinate')
    ax.set_title(f"{filename}\nDetected and Added Circles")
    ax.legend()
    ax.grid(True)

    return fig

def plot_final_register(result):
    brightfield_image = result.brightfield_image
    added = result.added
    removed = result.removed
    circles = result.snaked_circles
    filename = os.path.basename(result.file_path)

    # Determine the scaling factor
    max_dimension = 800
    h, w = brightfield_image.shape[:2]
    scale = min(max_dimension / h, max_dimension / w)

    # Resize the image to fit within the 800x800 box
    resized_image = cv2.resize(brightfield_image, (int(w * scale), int(h * scale)))

    # Create a figure and axis with matplotlib
    fig, ax = plt.subplots()
    ax.imshow(resized_image, cmap='gray')

    # Draw all circles in green
    for i in range(0,len(circles[0])):
        x = circles[0][i]
        y = circles[1][i]
        r = circles[2][i]
        col = circles[3][i]
        row = circles[4][i]
        index = circles[5][i]


        # Scale the circle positions and radius
        x = int(x * scale)
        y = int(y * scale)
        r = int(r * scale)

        # Draw the circle as a patch
        circle_patch = patches.Circle((x, y), r, edgecolor='green', facecolor='none', linewidth=2)
        ax.add_patch(circle_patch)

        # Annotate with row and column indices
        ax.text(x, y - r - 5, f"{index}", color='black', fontsize=8, ha='center')

    # Draw the removed circles in red with 50% transparency
    for x, y, r in zip(removed[0], removed[1], removed[2]):
        x = int(x * scale)
        y = int(y * scale)
        r = int(r * scale)

        circle_patch = patches.Circle((x, y), r, edgecolor='red', facecolor='none', linewidth=2, alpha=0.5)
        ax.add_patch(circle_patch)

    # Draw the added circles in blue
    for x, y, r in added:
        x = int(x * scale)
        y = int(y * scale)
        r = int(r * scale)

        circle_patch = patches.Circle((x, y), r, edgecolor='blue', facecolor='none', linewidth=2)
        ax.add_patch(circle_patch)

    # Set the aspect of the plot to be equal
    ax.set_aspect('equal')
    # Add grid lines
    ax.grid(True)
    # Display the result with interactive zoom
    ax.set_title(f"{filename}\nFinal Circle Registration")

    return fig

def plot_final_single_circle(result):
    brightfield_image = result.brightfield_image
    circle = result.single_circle
    filename = os.path.basename(result.file_path)

    # Determine the scaling factor
    max_dimension = 800
    h, w = brightfield_image.shape[:2]
    scale = min(max_dimension / h, max_dimension / w)

    # Resize the image to fit within the 800x800 box
    resized_image = cv2.resize(brightfield_image, (int(w * scale), int(h * scale)))

    # Create a figure and axis with matplotlib
    fig, ax = plt.subplots()
    ax.imshow(resized_image, cmap='gray')

        # Draw all circles in green
    x, y, r = circle

    x = int(x * scale)
    y = int(y * scale)
    r = int(r * scale)

    # Draw the circle as a patch
    circle_patch = patches.Circle((x, y), r, edgecolor='green', facecolor='none', linewidth=2)
    ax.add_patch(circle_patch)

    # Set the aspect of the plot to be equal
    ax.set_aspect('equal')

    # Add grid lines
    ax.grid(True)

    # Display the result with interactive zoom
    ax.set_title(f"{filename}\nFinal Circle Registration")

    return fig

def plot_fluorescence_vs_index(result):

    fluorescence_averages = result.fluorescence_averages
    num_channels = len(fluorescence_averages)
    filename = os.path.basename(result.file_path)

    fig, axes = plt.subplots(num_channels, 1, figsize=(12,10))

    # Handle single channel case
    if num_channels == 1:
        axes = [axes]

    for channel, averages in enumerate(fluorescence_averages):
        indices = averages[:,0]
        values = averages[:,1]

        axes[channel].plot(indices, values, 'go', label=f"Channel {channel+1}")

        axes[channel].set_xlabel("Snake Index")
        axes[channel].set_ylabel("Corrected Average Fluorescence")
        axes[channel].set_title(f"{filename}\nCorrected Average Fluorescence vs Snake Index - Channel {channel+1}")

        axes[channel].legend()
        axes[channel].grid(True)

    plt.tight_layout()

    return fig

def create_single_electrode_plot(result):

    averages = result.avg_values
    filename = os.path.basename(result.file_path)

    channels = [x[0] for x in averages]
    values = [x[1] for x in averages]

    fig, ax = plt.subplots(figsize=(6,5))

    bars = ax.bar(channels, values)

    for bar, value in zip(bars, values):
        ax.text(
            bar.get_x() + bar.get_width()/2,
            bar.get_height(),
            f"{value:.2f}",
            ha="center",
            va="bottom"
        )

    ax.set_xlabel("Channel")
    ax.set_ylabel("Corrected Average Fluorescence")
    ax.set_title(f"{filename}\nSingle Electrode Fluorescence")

    ax.grid(True, axis="y")

    plt.tight_layout()

    return fig

def plot_roi_bck(result, config):
    brightfield_image = result.brightfield_image
    circles = result.snaked_circles
    filename = os.path.basename(result.file_path)

    roi_inner = config["roi_inner"]
    roi_outer = config["roi_outer"]
    bckg_inner = config["bckg_inner"]
    bckg_outer = config["bckg_outer"]

    max_dimension = 800
    h, w = brightfield_image.shape[:2]
    scale = min(max_dimension / h, max_dimension / w)

    resized_image = cv2.resize(brightfield_image, (int(w * scale), int(h * scale)))


    fig, ax = plt.subplots(figsize=(10,8))

    ax.imshow(resized_image, cmap="gray")

    # loop through circles
    for i in range(len(circles[0])):

        x = circles[0][i]
        y = circles[1][i]
        r = circles[2][i]
        index = circles[5][i]

        # actual radii
        roi_outer_radius = r + roi_outer
        roi_inner_radius = r + roi_inner

        bckg_outer_radius = r + bckg_outer
        bckg_inner_radius = r + bckg_inner

        # scale for display
        xs = x * scale
        ys = y * scale

        # ROI outer
        ax.add_patch(patches.Circle((xs,ys), roi_outer_radius*scale, edgecolor="green", facecolor="none", linewidth=1))

        # ROI inner
        ax.add_patch(patches.Circle((xs,ys),roi_inner_radius*scale,edgecolor="green",facecolor="none",linestyle="--",linewidth=1))

        # background outer
        ax.add_patch(patches.Circle((xs,ys),bckg_outer_radius*scale,edgecolor="red",facecolor="none",linewidth=1))

        # background inner
        ax.add_patch(patches.Circle((xs,ys),bckg_inner_radius*scale,edgecolor="red",facecolor="none",linestyle="--",linewidth=1))
        ax.text(xs,ys,str(index),color="red", fontsize=8)

    ax.set_aspect("equal")
    ax.grid(True)
    ax.set_title(f"{filename}\nROI and Background Regions")

    return fig

def plot_roi_bck_single_electrode(result, config):
    brightfield_image = result.brightfield_image
    circles = result.single_circle
    filename = os.path.basename(result.file_path)

    roi_inner = config["roi_inner"]
    roi_outer = config["roi_outer"]
    bckg_inner = config["bckg_inner"]
    bckg_outer = config["bckg_outer"]

    max_dimension = 800
    h, w = brightfield_image.shape[:2]
    scale = min(max_dimension / h, max_dimension / w)

    resized_image = cv2.resize(brightfield_image, (int(w * scale), int(h * scale)))


    fig, ax = plt.subplots(figsize=(10,8))

    ax.imshow(resized_image, cmap="gray")

    x = circles[0]
    y = circles[1]
    r = circles[2]

    # actual radii
    roi_outer_radius = r + roi_outer
    roi_inner_radius = r + roi_inner

    bckg_outer_radius = r + bckg_outer
    bckg_inner_radius = r + bckg_inner

    # scale for display
    xs = x * scale
    ys = y * scale

    # ROI outer
    ax.add_patch(patches.Circle((xs,ys), roi_outer_radius*scale, edgecolor="green", facecolor="none", linewidth=1))

    # ROI inner
    ax.add_patch(patches.Circle((xs,ys),roi_inner_radius*scale,edgecolor="green",facecolor="none",linestyle="--",linewidth=1))

    # background outer
    ax.add_patch(patches.Circle((xs,ys),bckg_outer_radius*scale,edgecolor="red",facecolor="none",linewidth=1))

    # background inner
    ax.add_patch(patches.Circle((xs,ys),bckg_inner_radius*scale,edgecolor="red",facecolor="none",linestyle="--",linewidth=1))

    ax.set_aspect("equal")
    ax.grid(True)

    ax.set_title(f"{filename}\nROI and Background Regions")

    return fig

def fancy_plot(result):
    plot_data = result.fancy_plot_data
    filename = os.path.basename(result.file_path)

    fig = plt.figure(figsize=(12, 10))
    fig.suptitle(filename, fontsize = 14, fontweight = "bold")

    num_channels = len(plot_data)

    for channel_data in plot_data:

        channel = channel_data["channel"]
        indices = channel_data["indices"]
        values = channel_data["values"]
        smoothed_values = channel_data["smoothed_values"]
        outlier_flags = channel_data["outlier_flags"]
        peaks = channel_data["peaks"]
        avg_value = channel_data["avg_value"]

        plt.subplot(num_channels, 1, channel)

        # Plot the original data
        for i, (index, value) in enumerate(zip(indices, values)):
            if outlier_flags[i]:
                plt.plot(index, value, "yo", alpha=0.5)
            else:
                plt.plot(index, value, "go", alpha=0.5)

        # Plot smoothed curve
        plt.plot(
            indices,
            smoothed_values,
            "b-",
            label=f"Smoothed Channel {channel}"
        )

        # Plot peaks
        plt.plot(
            indices[peaks],
            smoothed_values[peaks],
            "ro",
            alpha=0.5,
            label="Peaks"
        )

        # Label peaks
        for peak in peaks:
            plt.text(
                indices[peak],
                smoothed_values[peak],
                f"{indices[peak]}\n{smoothed_values[peak]:.2f}",
                fontsize=8,
                ha="right",
                color="red"
            )

        # Average line
        plt.axhline(
            y=avg_value,
            color="black",
            linestyle="--",
            linewidth=1,
            label="Average"
        )

        plt.text(
            indices[-1],
            avg_value,
            f"Avg: {avg_value:.2f}",
            fontsize=8,
            color="black",
            ha="right",
            va="bottom"
        )

        plt.xlabel("Snake Index")
        plt.ylabel("Corrected Average Fluorescence")
        plt.title(f"Corrected Average Fluorescence - Channel {channel}")
        plt.legend()

    plt.tight_layout(rect=[0,0,1,0.96])

    return fig

def enhance_fluorescence(image, gamma=0.5):

    image = image.astype(np.float32)

    # contrast stretch
    low = np.percentile(image, 1)
    high = np.percentile(image, 99)

    image = np.clip(image, low, high)

    image = (image - low) / (high - low)

    # brighten dim signals
    image = image ** gamma

    return image

def plot_roi_background(result):

    plot_data = result.roi_plot_data
    filename = os.path.basename(result.file_path)
    figures = []

    for channel_data in plot_data:

        display_image = enhance_fluorescence(channel_data["image"])
        channel = channel_data["channel"]
        circles = channel_data["circles"]

        fig, ax = plt.subplots(figsize=(8,8))
        ax.imshow(display_image, cmap="gray")

        for circle in circles:
            x = circle["x"]
            y = circle["y"]
            roi_outer = circle["roi_outer"]
            roi_inner = circle["roi_inner"]
            background_outer = circle["background_outer"]
            background_inner = circle["background_inner"]
            index = circle["index"]

            # ROI green
            ax.add_patch(patches.Circle( (x,y), roi_outer, edgecolor="lime", facecolor="none", linewidth=1))
            ax.add_patch(patches.Circle((x,y), roi_inner, edgecolor="lime", facecolor="none", linewidth=1))

            # Background red
            ax.add_patch(patches.Circle((x,y), background_outer, edgecolor="red", facecolor="none", linewidth=1))
            ax.add_patch(patches.Circle((x,y), background_inner, edgecolor="red", facecolor="none", linewidth=1))
            ax.text(x, y, str(index),color="yellow",fontsize=10)

        # Display the result
        ax.set_title(f"{filename}\nChannel {channel}: ROI and Background Overlay")
        ax.axis("off")
        figures.append(fig)

    return figures

class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tip = None

        widget.bind("<Enter>", self.show)
        widget.bind("<Leave>", self.hide)

    def show(self, event=None):
        if self.tip:
            return

        x = self.widget.winfo_rootx() + 20
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 5

        self.tip = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")

        label = tk.Label(
            tw,
            text=self.text,
            justify="left",
            background="#ffffe0",
            relief="solid",
            borderwidth=1,
            padx=5,
            pady=3,
            wraplength=300
        )

        label.pack()

    def hide(self, event=None):
        if self.tip:
            self.tip.destroy()
            self.tip = None

class App:
    def __init__(self, root=None, processor=None):
        self.root = root or tk.Tk()
        self.processor = processor
        self.root.title("GridQuant")
        self.root.minsize(650, 650)
        self.root.attributes('-topmost', True)
        width = self.root.winfo_reqwidth()
        height = self.root.winfo_reqheight()
        self.root.geometry(f"{width}x{height}")
        self.cancel_event = threading.Event()
        self.processing = False
        self.plot_queue = []

        try:
            self.presets = load_presets()
            last_preset = self.load_last_preset()

            if last_preset in self.presets:
                default_preset = last_preset
            elif self.presets:
                default_preset = next(iter(self.presets))
            else: 
                default_preset = ""
        except (OSError, ValueError, json.JSONDecodeError) as error:
            messagebox.showerror("Preset error", str(error), parent=self.root)
            self.presets = {}

        ################################
        # Variables #
        ################################

        self.path = tk.StringVar()        
        self.status = tk.StringVar(value="Choose a file or folder.")
        self.progress_text = tk.StringVar(value="")
        self.preset_name = tk.StringVar(value=default_preset)
        self.debug = tk.BooleanVar(value=False)
        self.show_final_registration = tk.BooleanVar(value=False)
        self.show_if_images = tk.BooleanVar(value=False)
        self.show_plots = tk.BooleanVar(value=False)
        self.save_results_var = tk.BooleanVar(value=True)

        ################################
        # Variables to Add or Delete Preset #
        ################################

        self.new_name = tk.StringVar()
        self.new_is_array = tk.BooleanVar()
        self.new_min_diameter = tk.StringVar()
        self.new_max_diameter = tk.StringVar()
        self.new_roi_inner = tk.StringVar()
        self.new_roi_outer = tk.StringVar()
        self.new_bckg_inner = tk.StringVar()
        self.new_bckg_outer = tk.StringVar()
        self.new_moving_avg = tk.StringVar()
        self.new_p1 = tk.StringVar()
        self.new_p2 = tk.StringVar()
        self.new_dp = tk.StringVar()
        self.manage_preset = tk.StringVar()

        ################################
        # Notebook #
        ################################

        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)

        self.process_tab = ttk.Frame(self.notebook, padding=18)
        self.preset_tab = ttk.Frame(self.notebook, padding=18)
        self.help_tab = ttk.Frame(self.notebook, padding=18)

        self.notebook.add(self.process_tab, text="Image Processing")
        self.notebook.add(self.preset_tab, text="Create Preset")
        self.notebook.add(self.help_tab, text="Help")

        ################################
        # Build Tabs #
        ################################

        self.build_process_tab()
        self.build_preset_tab()
        self.build_help_tab()

        # Automatically size window to fit widgets
        self.root.update_idletasks()
        self.root.geometry(f"{self.root.winfo_reqwidth()}x{self.root.winfo_reqheight()}")
        self.root.minsize(650, 600)

    def build_process_tab(self):

        frame = ttk.Frame(self.process_tab, padding=20)
        frame.pack(fill="both", expand=True)

        frame.columnconfigure(1, weight=1)

        ###################################################
        # Title
        ###################################################

        ttk.Label(
            frame,
            text="GridQuant",
            font=("Segoe UI", 18, "bold")
        ).grid(
            row=0,
            column=0,
            columnspan=3,
            sticky="w",
            pady=(0,20)
        )

        ###################################################
        # Input
        ###################################################

        ttk.Label(
            frame,
            text="Input"
        ).grid(
            row=1,
            column=0,
            sticky="w",
            pady=5
        )

        ttk.Entry(
            frame,
            textvariable=self.path,
            state="readonly"
        ).grid(
            row=1,
            column=1,
            sticky="ew",
            padx=10
        )

        ttk.Button(
            frame,
            text="Browse File",
            command=self.choose_file
        ).grid(
            row=1,
            column=2,
            padx=5
        )

        ttk.Button(
            frame,
            text="Browse Folder",
            command=self.choose_folder
        ).grid(
            row=1,
            column=3,
            padx=5
        )

        ###################################################
        # Preset
        ###################################################

        ttk.Label(
            frame,
            text="Preset"
        ).grid(
            row=3,
            column=0,
            sticky="w",
            pady=15
        )

        self.preset_combo = ttk.Combobox(
            frame,
            textvariable=self.preset_name,
            values=sorted(self.presets.keys()),
            state="readonly"
        )

        
        self.preset_combo.grid(
            row=3,
            column=1,
            columnspan=2,
            sticky="ew",
            padx=10
        )

        ###################################################
        # Options
        ###################################################

        options = ttk.LabelFrame(
            frame,
            text="Options",
            padding=10
        )

        options.grid(
            row=4,
            column=0,
            columnspan=3,
            sticky="ew",
            pady=20
        )

        options.columnconfigure(0, weight=1)
        options.columnconfigure(1, weight=1)

        ttk.Checkbutton(options, text="Save results", variable=self.save_results_var).grid(row=0, column=0, sticky="w")

        ttk.Checkbutton(
            options,
            text="Debug",
            variable=self.debug
        ).grid(row=1,column=0,sticky="w")

        ttk.Checkbutton(
            options,
            text="Show Final Registration*",
            variable=self.show_final_registration
        ).grid(row=0,column=1,sticky="w")

        ttk.Checkbutton(
            options,
            text="Show I.F. Images",
            variable=self.show_if_images
        ).grid(row=0,column=2,sticky="w")

        ttk.Checkbutton(
            options,
            text="Show Plots",
            variable=self.show_plots
        ).grid(row=1,column=1,sticky="w")

        ###################################################
        # Buttons
        ###################################################

        button_frame = ttk.Frame(frame)

        button_frame.grid(
            row=5,
            column=0,
            columnspan=3,
            pady=(5,15)
        )

        self.start_button = ttk.Button(
            button_frame,
            text="Start Processing",
            command=self.submit
        )

        self.start_button.pack(
            side="left",
            padx=5
        )

        self.cancel_button = ttk.Button(
            button_frame,
            text="Cancel",
            command=self.cancel_processing,
            state="disabled"
        )

        self.cancel_button.pack(
            side="left",
            padx=5
        )

        ###################################################
        # Progress
        ###################################################

        ttk.Label(
            frame,
            text="Progress"
        ).grid(
            row=6,
            column=0,
            sticky="w"
        )

        self.progress = ttk.Progressbar(
            frame,
            orient="horizontal",
            mode="determinate",
            maximum=100
        )

        self.progress.grid(
            row=7,
            column=0,
            columnspan=3,
            sticky="ew",
            pady=5
        )

        ###################################################
        # Current file
        ###################################################

        ttk.Label(
            frame,
            textvariable=self.progress_text,
            foreground="blue"
        ).grid(
            row=8,
            column=0,
            columnspan=3,
            sticky="w",
            pady=(8,0)
        )

        ###################################################
        # Status
        ###################################################

        ttk.Label(
            frame,
            textvariable=self.status,
            wraplength=580
        ).grid(
            row=9,
            column=0,
            columnspan=3,
            sticky="w",
            pady=(10,0)
        )

        ###################################################
        # Processing log window
        ###################################################

        ttk.Label(
            frame,
            text="Processing Log"
        ).grid(
            row=10,
            column=0,
            sticky="w",
            pady=(10,0)
        )

        self.log_box = tk.Text(
            frame,
            height=10,
            width=70,
            state="disabled"
        )

        self.log_box.grid(
            row=11,
            column=0,
            columnspan=3,
            sticky="nsew",
            pady=5
        )

        ttk.Button(
            frame,
            text="Clear Log",
            command=self.clear_log
        ).grid(
            row=11,
            column=3,
            padx=10,
            sticky="w"
        )

        frame.rowconfigure(11, weight=1)
        frame.columnconfigure(0, weight=1)

    def build_preset_tab(self):

        frame = ttk.Frame(
            self.preset_tab,
            padding=20
        )

        frame.pack(
            fill="both",
            expand=True
        )

        frame.columnconfigure(
            1,
            weight=1
        )

        ###################################################
        # Title
        ###################################################

        ttk.Label(
            frame,
            text="Create New Preset",
            font=("Segoe UI",16,"bold")
        ).grid(
            row=0,
            column=0,
            columnspan=2,
            sticky="w",
            pady=(0,20)
        )

        ###################################################
        # Copy Existing Preset
        ###################################################

        ttk.Label(
            frame,
            text="Copy Existing Preset",
        ).grid(
            row=2,
            column=0,
            sticky="w",
        )

        self.copy_preset_name = tk.StringVar()

        self.copy_combo = ttk.Combobox(
            frame,
            textvariable=self.copy_preset_name,
            values=list(self.presets.keys()),
            state="readonly",
            width=30
        )

        self.copy_combo.grid(
            row=2,
            column=1,
            sticky="ew",
            padx=5
        )

        self.copy_combo.bind(
            "<<ComboboxSelected>>",
            lambda event: self.load_preset_values()
        )

        ###################################################
        # Preset name
        ###################################################

        ttk.Label(
            frame,
            text="Preset Name"
        ).grid(
            row=4,
            column=0,
            sticky="w",
            pady=5
        )

        ttk.Entry(
            frame,
            textvariable=self.new_name
        ).grid(
            row=4,
            column=1,
            sticky="ew",
            padx=10
        )

        ###################################################
        # Array checkbox
        ###################################################

        ttk.Checkbutton(
            frame,
            text="Array",
            variable=self.new_is_array
        ).grid(
            row=3,
            column=0,
            columnspan=2,
            sticky="w",
            pady=5
        )

        ###################################################
        # Preset fields
        ###################################################

        fields = [

            ("Minimum Diameter", self.new_min_diameter),
            ("Maximum Diameter", self.new_max_diameter),
            ("ROI Inner", self.new_roi_inner),
            ("ROI Outer", self.new_roi_outer),
            ("Background Inner", self.new_bckg_inner),
            ("Background Outer", self.new_bckg_outer),
            ("Moving Average", self.new_moving_avg),
            ("P1", self.new_p1),
            ("P2", self.new_p2),
            ("DP", self.new_dp),
        ]


        for row,(label,var) in enumerate(fields, start=5):

            ttk.Label(
                frame,
                text=label
            ).grid(
                row=row,
                column=0,
                sticky="w",
                pady=3
            )


            ttk.Entry(
                frame,
                textvariable=var
            ).grid(
                row=row,
                column=1,
                sticky="ew",
                padx=10
            )

        ###################################################
        # Save button
        ###################################################

        ttk.Button(
            frame,
            text="Save Preset",
            command=self.save_preset
        ).grid(
            row=15,
            column=0,
            columnspan=2,
            pady=20
        )

        ###################################################
        # Delete Preset
        ###################################################

        ttk.Separator(frame).grid(
            row=16,
            column=0,
            columnspan=3,
            sticky="ew",
            pady=5
        )

        ttk.Label(frame, text="Delete Presets", font=("Segoe UI", 16, "bold")).grid(
            row=17,
            column=0,
            columnspan=3,
            sticky="w",
            pady=(0,10)
        )

        self.delete_preset_name = tk.StringVar(
            value=next(iter(self.presets), "")
        )

        ttk.Label(frame, text="Preset").grid(
            row=18,
            column=0,
            sticky="w"
        )

        self.delete_combo = ttk.Combobox(
            frame,
            textvariable=self.delete_preset_name,
            values=list(self.presets.keys()),
            state="readonly",
            width=30
        )

        self.delete_combo.grid(
            row=18,
            column=1,
            sticky="ew",
            padx=5
        )

        ttk.Button(
            frame,
            text="Delete",
            command=self.delete_preset
        ).grid(
            row=18,
            column=2,
            padx=5
        )

    def build_help_tab(self):

        frame = ttk.Frame(self.help_tab)
        frame.pack(fill="both", expand=True)

        scrollbar = ttk.Scrollbar(frame)
        scrollbar.pack(side="right", fill="y")

        help_box = tk.Text(
            frame,
            wrap="word",
            yscrollcommand=scrollbar.set,
            font=("Arial", 11),
            padx=10,
            pady=10)

        help_box.pack(side="left", fill="both", expand=True)
        help_box.tag_configure("title", font=("Arial", 18, "bold")  )

        help_box.tag_configure("header", font=("Arial", 14, "bold"))

        help_box.tag_configure("subheader", font=("Arial", 12, "bold"))

        help_box.tag_configure("bold", font=("Arial", 11, "bold"))

        help_box.tag_configure("bullet", lmargin1=20, lmargin2=40)

        help_box.tag_configure("normal", font=("Arial", 11))
        scrollbar.config(command=help_box.yview)

        help_text = self.load_help_file()
        self.insert_markdown(help_box, help_text)
        help_box.config(state="disabled")

    def load_help_file(self):

        help_path = os.path.join(os.path.dirname(__file__), "help.md")

        try:
            with open(help_path, "r", encoding="utf-8") as file:
                return file.read()

        except FileNotFoundError:
            return "# Help file missing\n\nCould not find help.md"

    def insert_markdown(self, widget, text):

        lines = text.split("\n")

        for line in lines:

            # Main title #
            if line.startswith("# "):
                widget.insert("end", line[2:] + "\n", "title")

            # Header ##
            elif line.startswith("## "):
                widget.insert("end", "\n" + line[3:] + "\n", "header")

            # Header ###
            elif line.startswith("### "):
                widget.insert("end","\n" + line[4:] + "\n","subheader")

            # Bullet list
            elif line.startswith("- "):

                widget.insert("end", "• " + line[2:] + "\n",  "bullet")

            # Blank lines
            elif line.strip() == "":
                widget.insert("end", "\n")

            # Bold text **example**
            else:
                parts = re.split(r"(\*\*.*?\*\*)", line)

                for part in parts:
                    if part.startswith("**") and part.endswith("**"):
                        widget.insert("end", part[2:-2], "bold")

                    else:
                        widget.insert("end", part, "normal")
                widget.insert("end", "\n")

    def load_last_preset(self):
        if os.path.exists(LAST_CONFIG_FILE):
            try:
                with open(LAST_CONFIG_FILE, "r") as f:
                    data = json.load(f)
                    return data.get("last_preset", None)
            except Exception:
                pass

        return ""

    def save_last_preset(self, preset_name):
        try:
            with open(LAST_CONFIG_FILE, "w") as f:
                json.dump({"last_preset": preset_name}, f, indent=4)
        except Exception:
            pass

    def save_preset(self):

        name = self.new_name.get().strip()
        if not name:
            messagebox.showwarning("Missing Name", "Please enter a preset name.", parent=self.root)
            return

        if name in self.presets:

            overwrite = messagebox.askyesno("Overwrite preset?", f"{name} already exists.\nOverwrite it?", parent=self.root)

            if not overwrite:
                return

        try:

            preset = {
                "is_array":
                    self.new_is_array.get(),
                "min_diameter":
                    int(self.new_min_diameter.get()),
                "max_diameter":
                    int(self.new_max_diameter.get()),
                "roi_inner":
                    int(self.new_roi_inner.get()),
                "roi_outer":
                    int(self.new_roi_outer.get()),
                "bckg_inner":
                    int(self.new_bckg_inner.get()),
                "bckg_outer":
                    int(self.new_bckg_outer.get()),
                "moving_avg_n":
                    int(self.new_moving_avg.get()),
                "p1":
                    float(self.new_p1.get()),
                "p2":
                    float(self.new_p2.get()),
                "dp":
                    float(self.new_dp.get()),
            }

        except ValueError:

            messagebox.showerror("Invalid Input", "All fields must contain valid numbers.", parent=self.root)

            return

        ###################################################
        # Update JSON
        ###################################################

        self.presets[name] = preset

        path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "presets.json")

        with open(path, "w", encoding="utf-8") as f:

            json.dump(self.presets, f, indent=4)

        self.delete_combo["values"] = list(self.presets.keys())
        self.preset_combo["values"] = list(self.presets.keys())
        self.copy_combo["values"] = list(self.presets.keys())

        ###################################################
        # Refresh dropdown
        ###################################################

        self.preset_combo["values"] = sorted(self.presets.keys())

        new_values = list(self.presets.keys())

        self.preset_combo["values"] = new_values
        self.delete_combo["values"] = new_values
        self.copy_combo["values"] = new_values

        self.preset_name.set(name)
        self.delete_preset_name.set(name)
        self.copy_preset_name.set(name)

        ###################################################
        # Clear form
        ###################################################

        self.new_name.set("")

        self.new_min_diameter.set("")
        self.new_max_diameter.set("")

        self.new_roi_inner.set("")
        self.new_roi_outer.set("")

        self.new_bckg_inner.set("")
        self.new_bckg_outer.set("")

        self.new_moving_avg.set("")

        self.new_p1.set("")
        self.new_p2.set("")
        self.new_dp.set("")


        messagebox.showinfo("Saved", f"Preset '{name}' saved.")

    def choose_file(self):
        selected = filedialog.askopenfilename(
            parent=self.root,
            title="Select CZI image",
            filetypes=[("CZI files", "*.czi"), ("All files", "*.*")],
        )
        if selected:
            self.path.set(selected)

    def choose_folder(self):
        selected = filedialog.askdirectory(parent=self.root, title="Select image folder")
        if selected:
            self.path.set(selected)

    def update_progress(self, current, total, filename):

        self.progress["maximum"] = total
        self.progress["value"] = current

        self.status.set(f"Processing {current}/{total}\n{os.path.basename(filename)}")

        self.root.update_idletasks()

    def cancel_processing(self):

        if self.processing:
            self.cancel_event.set()
            self.status.set("Cancel requested...")

    def submit(self):
        if self.processing:
            return
        
        selected_path = self.path.get()
        selected_preset = self.preset_name.get()

        if not selected_path or not os.path.exists(selected_path):
            messagebox.showwarning("Input required", "Choose a CZI file or folder first.", parent=self.root)
            return
        if selected_preset not in self.presets:
            messagebox.showwarning("Preset required", "Choose a preset from presets.json.", parent=self.root)
            return

        runtime_settings = {
            "debug": self.debug.get(),
            "show_final_registration": self.show_final_registration.get(),
            "show_if_images": self.show_if_images.get(),
            "show_plots": self.show_plots.get(),
        }
        self.status.set("Processing... see the terminal for progress.")
        self.root.update_idletasks()
        try:
            if self.processor is None:
                from gridquant_V2_0 import run_pipeline
            else:
                run_pipeline = self.processor.run_pipeline

            config = {
                "path": selected_path,
                # preset values
                **self.presets[selected_preset],
                # GUI options
                **runtime_settings,}


            config = {
                "path": selected_path,
                **self.presets[selected_preset],
                **runtime_settings,
                "preset_name": selected_preset,
            }

            self.processing = True
            self.cancel_event.clear()

            self.start_button.config(state="disabled")
            self.cancel_button.config(state="normal")

            thread = threading.Thread(
                target=self.run_processing,
                args=(run_pipeline, config, selected_path),
                daemon=True
            )

            thread.start()

        except Exception as error:
            self.status.set("Processing failed.")
            messagebox.showerror("Processing error", str(error), parent=self.root)

    def update_log(self, message):

        def append():

            self.log_box.config(state="normal")
            self.log_box.insert(tk.END, message + "\n")
            self.log_box.see(tk.END)
            self.log_box.config(state="disabled")

        self.root.after(0, append)

    def clear_log(self):

        def clear():

            self.log_box.config(state="normal")
            self.log_box.delete("1.0", tk.END)

            self.log_box.config(state="disabled")

        self.root.after(0, clear)

    def add_log_section(self, title):

        separator = "=" * 50

        self.update_log("")
        self.update_log(separator)
        self.update_log(title)
        self.update_log(separator)
    
    def update_progress(self, current, total, filename):

        self.root.after(
            0,
            self.update_progress,
            current,
            total,
            filename
        )

    def display_debug_plots(self, results, config):
        for result in results:
            debug_plots = []

            if config["is_array"]:
                debug_plots.append(fancy_plot(result))

            for debug_plot in debug_plots:
                if debug_plot is not None:
                    debug_plot.show()

    def display_figures(self, results, config):

        for result in results: 
            figures = []
            figures.append(draw_circles_and_display(result))

            if config["is_array"]:
                figures.append(plot_outliers(result))
                figures.append(plot_missing(result))
            
            for fig in figures:
                if fig is not None: 
                    fig.show()

    def display_final_reg(self, results, config):
        for result in results:
            final_plots = []

            if config["is_array"]: 
                final_plots.append(plot_final_register(result))
                final_plots.append(plot_roi_bck(result, config))
            else:
                final_plots.append(plot_final_single_circle(result))
                final_plots.append(plot_roi_bck_single_electrode(result, config))

            for plot in final_plots:
                if plot is not None:
                    plt.show()
                    
    def save_roi_background_plots(self, results, selected_path, config):

        if not results:
            return None

        # Same directory where the Excel file is being saved
        output_dir = (
            selected_path
            if os.path.isdir(selected_path)
            else os.path.dirname(selected_path)
        )

        # Create folder
        roi_dir = os.path.join(
            output_dir,
            "Final Plots"
        )

        os.makedirs(roi_dir, exist_ok=True)

        for result in results:

            # Get filename without extension
            filename = os.path.splitext(
                os.path.basename(result.file_path)
            )[0]

            # Create the appropriate ROI/background plot
            if config["is_array"]:
                plot = plot_roi_bck(result, config)
            else:
                plot = plot_roi_bck_single_electrode(result, config)

            if plot is None:
                continue

            # Save image
            output_file = os.path.join(
                roi_dir,
                f"{filename}_ROI_background.png"
            )

            plot.savefig(
                output_file,
                dpi=300,
                bbox_inches="tight"
            )

            # Don't display it
            plt.close(plot)

        return roi_dir

    def display_if_plots(self, results, config):

        if_plots = []

        for result in results:
            if_plots.extend(plot_roi_background(result))

            if config["is_array"]:
                if_plots.append(plot_fluorescence_vs_index(result))

            else:
                if_plots.append(create_single_electrode_plot(result))

        for if_plot in if_plots:
            if if_plot is not None:
                if_plot.show()

    def update_progress(self, current, total, filename):

        self.progress["maximum"] = total
        self.progress["value"] = current
        self.status.set(f"Processing {current}/{total}: {os.path.basename(filename)}")    

    def save_results(self, results, selected_path, config):
        if not results:
            return None

        # ---------------------------------------------------------
        # Determine output location
        # ---------------------------------------------------------
        output_dir = (
            selected_path
            if os.path.isdir(selected_path)
            else os.path.dirname(selected_path)
        )

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        output_file = os.path.join(
            output_dir,
            f"summary_{timestamp}.xlsx"
        )

        # ---------------------------------------------------------
        # Create workbook
        # ---------------------------------------------------------
        wb = Workbook()

        # =========================================================
        # SHEET 1 - Summary
        # =========================================================
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Version
        ws_summary.append([
            "GridQuant Version",
            GRIDQUANT_VERSION
        ])

        # Analysis date
        ws_summary.append([
            "Analysis Date",
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ])

        ws_summary.append([])

        # Configuration
        ws_summary.append([
            "Configuration",
            config.get("preset_name")
        ])

        for key, value in config.items():
            ws_summary.append([key, value])

        ws_summary.append([])

        # Existing summary header
        ws_summary.append([
            "File",
            "Channel",
            "Average Fluorescence"
        ])

        # Existing summary data
        for result in results:
            filename = os.path.basename(result.file_path)

            for channel, value in result.avg_values:
                ws_summary.append([
                    filename,
                    channel,
                    float(value)
                ])

        # =========================================================
        # SHEET 2 - Individual Electrode Data
        # =========================================================

        ws_electrodes = wb.create_sheet("Individual Electrode Data")

        ws_electrodes.append([
            "File",
            "Channel",
            "Electrode",
            "X",
            "Y",
            "R",
            "Corrected Fluorescence",
            "Outlier",
            "Cleaned Fluorescence"
        ])


        for result in results:

            filename = os.path.basename(result.file_path)

            # =====================================================
            # ARRAY DATA
            # =====================================================

            if result.fancy_plot_data is not None:

                # Get electrode coordinates from snaked_circles
                x_values = result.snaked_circles[0]
                y_values = result.snaked_circles[1]
                r_values = result.snaked_circles[2]
                electrode_indices = result.snaked_circles[5]

                # Create lookup:
                # electrode ID -> (x, y, r)
                circle_lookup = {}

                for x, y, r, idx in zip(
                    x_values,
                    y_values,
                    r_values,
                    electrode_indices
                ):
                    circle_lookup[int(idx)] = (
                        x,
                        y,
                        r
                    )

                # Loop through channels
                for channel_data in result.fancy_plot_data:

                    channel = channel_data["channel"]

                    channel_name = result.channel_names[
                        channel - 1
                    ]

                    indices = channel_data["indices"]
                    values = channel_data["values"]
                    cleaned_values = channel_data["cleaned_values"]
                    outlier_flags = channel_data["outlier_flags"]

                    # Loop through electrodes
                    for electrode_id, corrected, cleaned, is_outlier in zip(
                        indices,
                        values,
                        cleaned_values,
                        outlier_flags
                    ):

                        electrode_id = int(electrode_id)

                        # Get X, Y, R
                        x, y, r = circle_lookup[electrode_id]

                        ws_electrodes.append([
                            filename,
                            channel_name,
                            electrode_id,
                            float(x),
                            float(y),
                            float(r),
                            float(corrected),
                            "Yes" if is_outlier else "No",
                            float(cleaned)
                        ])


            # =====================================================
            # SINGLE ELECTRODE DATA
            # =====================================================

            else:

                # Convert named average values into a dictionary
                # Example:
                # {"EGFP": 1254.3, "DsRed": 542.2}
                avg_lookup = dict(result.avg_values)

                # roi_plot_data contains the X, Y, R information
                for channel_data in result.roi_plot_data:

                    channel = channel_data["channel"]

                    channel_name = result.channel_names[
                        channel - 1
                    ]

                    # There is only one electrode
                    circle = channel_data["circles"][0]

                    x = circle["x"]
                    y = circle["y"]
                    r = circle["r"]

                    # Get corrected fluorescence
                    corrected = avg_lookup[channel_name]

                    ws_electrodes.append([
                        filename,
                        channel_name,
                        1,
                        float(x),
                        float(y),
                        float(r),
                        float(corrected),
                        "No",
                        float(corrected)
                    ])

        # =========================================================
        # Formatting - Sheet 1
        # =========================================================

        for cell in ws_summary[1]:
            cell.font = Font(bold=True)

        # Find the summary header row
        for row in ws_summary.iter_rows():
            if row[0].value == "File" and row[1].value == "Channel":
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        "solid",
                        fgColor="D9EAF7"
                    )
                break

        # =========================================================
        # Formatting - Sheet 2
        # =========================================================

        for cell in ws_electrodes[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                "solid",
                fgColor="D9EAF7"
            )
            cell.alignment = Alignment(
                horizontal="center"
            )

        # Freeze header row
        ws_electrodes.freeze_panes = "A2"

        # Add autofilter
        ws_electrodes.auto_filter.ref = (
            ws_electrodes.dimensions
        )

        # =========================================================
        # Format numeric columns
        # =========================================================

        for row in ws_electrodes.iter_rows(
            min_row=2,
            min_col=4,
            max_col=5
        ):
            for cell in row:
                cell.number_format = "0.00"

        # =========================================================
        # Automatically size columns
        # =========================================================

        for ws in [ws_summary, ws_electrodes]:

            for column_cells in ws.columns:

                max_length = 0
                column_letter = get_column_letter(
                    column_cells[0].column
                )

                for cell in column_cells:
                    if cell.value is not None:
                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )

                ws.column_dimensions[
                    column_letter
                ].width = min(max_length + 2, 40)

        # ---------------------------------------------------------
        # Save workbook
        # ---------------------------------------------------------
        wb.save(output_file)

        return output_file

    def delete_preset(self):

        preset = self.delete_preset_name.get()

        if preset not in self.presets:
            messagebox.showwarning(
                "Delete Preset",
                "Please select a preset to delete.",
                parent=self.root
            )
            return

        confirm = messagebox.askyesno(
            "Delete Preset",
            f"Are you sure you want to permanently delete\n\n'{preset}'?",
            icon="warning",
            parent=self.root
        )

        if not confirm:
            return

        # Remove preset
        del self.presets[preset]

        # Save updated JSON
        with open("presets.json", "w", encoding="utf-8") as f:
            json.dump(self.presets, f, indent=4)

        # Refresh comboboxes
        values = list(self.presets.keys())

        self.preset_combo["values"] = values
        self.delete_combo["values"] = values
        self.copy_combo["values"] = values

        if values:
            self.preset_name.set(values[0])
            self.delete_preset_name.set(values[0])
            self.copy_preset_name.set(values[0])
        else:
            self.preset_name.set("")
            self.delete_preset_name.set("")
            self.copy_preset_name.set("")

        messagebox.showinfo(
            "Preset Deleted",
            f"'{preset}' has been deleted.",
            parent=self.root
        )

    def load_preset_values(self):

        name = self.copy_preset_name.get()

        if name not in self.presets:
            return

        preset = self.presets[name]

        self.new_is_array.set(preset["is_array"])
        self.new_min_diameter.set(preset["min_diameter"])
        self.new_max_diameter.set(preset["max_diameter"])
        self.new_roi_inner.set(preset["roi_inner"])
        self.new_roi_outer.set(preset["roi_outer"])
        self.new_bckg_inner.set(preset["bckg_inner"])
        self.new_bckg_outer.set(preset["bckg_outer"])
        self.new_moving_avg.set(preset["moving_avg_n"])
        self.new_p1.set(preset["p1"])
        self.new_p2.set(preset["p2"])
        self.new_dp.set(preset["dp"])

    def run_processing(self, run_pipeline, config, selected_path):

        try:
            self.save_last_preset(self.preset_name.get())
            results = run_pipeline(
                config,
                progress_callback=self.update_progress,
                log_callback=self.update_log,
                cancel_event=self.cancel_event
            )

            for result in results:

                self.update_log("")
                self.update_log("=" * 50)
                self.update_log("RESULTS")
                self.update_log("=" * 50)

                self.update_log(os.path.basename(result.file_path))

                for name, value in result.avg_values:
                    self.update_log(f"{name}: {value:.2f}")

            if config["debug"]:
                self.root.after(0, lambda: self.display_debug_plots(results, config))

            if config["show_plots"]:
                self.root.after(0, lambda: self.display_figures(results, config))

            if config["show_final_registration"]:
                self.root.after(0, lambda: self.display_final_reg(results, config))

            if config["show_if_images"]:
                self.root.after(0, lambda: self.display_if_plots(results, config))

            if self.save_results_var.get():
                output_file = self.save_results(results, selected_path, config)
                self.save_roi_background_plots(results, selected_path, config)

                self.root.after(
                    0,
                    lambda: messagebox.showinfo(
                        "Complete",
                        f"Summary saved to:\n{output_file}",
                        parent=self.root
                    )
                )
            else:
                self.root.after(0, lambda: self.status.set(f"Finished processing {len(results)} file(s). Results not saved to a CSV."))

        except Exception as error:

            traceback.print_exc()

            error_message = repr(error)

            self.root.after(
                0,
                lambda: messagebox.showerror(
                    "Processing error",
                    error_message,
                    parent=self.root
                )
            )

        finally:
            self.processing = False
            self.root.after(0, lambda: self.start_button.config(state="normal"))
            self.root.after(0, lambda: self.cancel_button.config(state="disabled"))
            self.progress["value"] = 0
                    
    def mainloop(self):
        self.root.mainloop()

if __name__ == "__main__":
    App().mainloop()
